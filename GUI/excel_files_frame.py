import customtkinter
from GUI.excel_files_window import ExcelFilesWindow
from GUI.frame import Frame
from Logic.data_file import DataClass
import openpyxl


class ExcelFilesFrame(Frame):
    def __init__(self, master, data: DataClass, **kwargs):
        super().__init__(master, **kwargs)

        self.rowconfigure(0, weight=1)
        self.columnconfigure((0, 1), weight=1)

        self.data = data

        self.create_xlsx_button = customtkinter.CTkButton(self, text="Create xlsx", command=self.create_excel_file,
                                                          state="disabled")
        self.create_xlsx_button.grid(row=0, column=0, padx=(5, 5), pady=(5, 5))
        self.interactive_elements.append(self.create_xlsx_button)

        self.info_button = customtkinter.CTkButton(self, text="Loaded xlsx", command=self.show_excel_files_info)
        self.info_button.grid(row=0, column=1, padx=(5, 5), pady=(5, 5))
        self.interactive_elements.append(self.info_button)

    def create_excel_file(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active

        for col, header in enumerate(self.data.headers):
            sheet.cell(row=1, column=col + 1).value = header

        row = 2
        for part_number, search_results in self.data.all_search_results:
            for search_result in sorted(search_results, key=lambda d: d["vendor"]):
                if (self.data.websites_to_search[search_result["vendor"].split()[0].capitalize()] and
                        self.data.conditions_to_search[self.data.get_condition(search_result["condition"])]):
                    for col, header in enumerate(self.data.headers):
                        sheet.cell(row=row, column=col + 1).value = search_result[header]
                    row += 1

        workbook.save(filename=customtkinter.filedialog.asksaveasfilename(defaultextension="xlsx",
                                                                          filetypes=[("Excel", "*.xlsx")]))

    def show_excel_files_info(self):
        ExcelFilesWindow(master=self.master, data=self.data).grab_set()
